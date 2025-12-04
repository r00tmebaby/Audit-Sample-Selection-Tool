"""Tests for Excel reporting outputs."""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from worker.src.models import (
    CleanedTransaction,
    DataQualityReport,
    SampleStatistics,
    SamplingParameters,
)
from worker.src.reporter import generate_reports


def _sample_quality() -> DataQualityReport:
    """Create sample quality report for testing."""
    return DataQualityReport(
        total_rows_raw=10,
        total_rows_cleaned=9,
        missing_transaction_id=1,
        missing_amount=1,
        missing_effective_date=0,
        missing_document_type=0,
        missing_description=0,
        invalid_amount_format=1,
        invalid_date_format=0,
        duplicate_transaction_ids=0,
        excluded_due_to_amount=1,
        excluded_due_to_balance=2,
        excluded_zero_amounts=3,
        notes="",
    )


def _sample_stats() -> SampleStatistics:
    """Create sample statistics for testing."""
    return SampleStatistics(
        population_size=100,
        population_balance_abs=1000.0,
        sampling_interval=50.0,
        high_value_count=5,
        random_sample_count=10,
        coverage_abs=400.0,
        coverage_percent=40.0,
        excluded_zero_amounts=3,
        excluded_due_to_balance=2,
    )


def _sample_params() -> SamplingParameters:
    """Create sample parameters for testing."""
    return SamplingParameters(
        tolerable_misstatement=500.0,
        expected_misstatement=100.0,
        assurance_factor=4.0,
        random_seed=7,
    )


def test_generate_reports_creates_excel(tmp_path: Path) -> None:
    """Report generation creates Excel workbook with required sheets."""
    sample = [
        CleanedTransaction(
            transaction_id="A",
            amount_signed=100.0,
            amount_abs=100.0,
            balance_category="credit",
            source_row_index=0,
            selection_type="High Value",
        )
    ]

    output_path = generate_reports(
        tmp_path,
        sample,
        _sample_quality(),
        _sample_stats(),
        _sample_params(),
        datetime.now(timezone.utc),
        "run-123",
    )

    assert output_path.exists()
    assert output_path.name == "sample_selection_output.xlsx"

    # Load and verify workbook structure
    workbook = load_workbook(output_path)

    assert "Population Summary" in workbook.sheetnames
    assert "Sample Selected" in workbook.sheetnames
    assert "Parameters Used" in workbook.sheetnames

    workbook.close()


def test_population_summary_content(tmp_path: Path) -> None:
    """Population Summary sheet contains expected metrics."""
    sample = []

    output_path = generate_reports(
        tmp_path,
        sample,
        _sample_quality(),
        _sample_stats(),
        _sample_params(),
        datetime.now(timezone.utc),
        "run-456",
    )

    workbook = load_workbook(output_path)
    sheet = workbook["Population Summary"]

    # Check headers
    assert sheet["A1"].value == "Metric"
    assert sheet["B1"].value == "Value"

    # Check some key metrics
    metrics = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:
            metrics[row[0]] = row[1]

    assert metrics["Total Population Value"] == 1000.0
    assert metrics["Number of Items"] == 100
    assert metrics["Random Seed Used"] == 7
    assert "Zero Excluded" in metrics["Data Quality Issues"]
    assert "Balance Excluded" in metrics["Data Quality Issues"]

    workbook.close()


def test_sample_selected_content(tmp_path: Path) -> None:
    """Sample Selected sheet contains transaction data."""
    sample = [
        CleanedTransaction(
            transaction_id="T001",
            amount_signed=150.0,
            amount_abs=150.0,
            balance_category="credit",
            source_row_index=0,
            selection_type="High Value",
        ),
        CleanedTransaction(
            transaction_id="T002",
            amount_signed=-50.0,
            amount_abs=50.0,
            balance_category="debit",
            source_row_index=1,
            selection_type="Random",
        ),
    ]

    output_path = generate_reports(
        tmp_path,
        sample,
        _sample_quality(),
        _sample_stats(),
        _sample_params(),
        datetime.now(timezone.utc),
        "run-789",
    )

    workbook = load_workbook(output_path)
    sheet = workbook["Sample Selected"]

    # Check headers exist (now at row 3 due to coverage banner and spacer)
    assert sheet["A3"].value == "Transaction ID"
    assert sheet["B3"].value == "Amount Signed"

    # Check data rows now start at row 4
    assert sheet["A4"].value == "T001"
    assert sheet["B4"].value == 150.0
    assert sheet["H4"].value == "High Value"

    assert sheet["A5"].value == "T002"
    assert sheet["B5"].value == -50.0
    assert sheet["H5"].value == "Random"

    workbook.close()


def test_parameters_used_content(tmp_path: Path) -> None:
    """Parameters Used sheet contains sampling parameters."""
    sample = []

    output_path = generate_reports(
        tmp_path,
        sample,
        _sample_quality(),
        _sample_stats(),
        _sample_params(),
        datetime.now(timezone.utc),
        "run-xyz",
    )

    workbook = load_workbook(output_path)
    sheet = workbook["Parameters Used"]

    # Check parameters
    params = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:
            params[row[0]] = row[1]

    assert params["Tolerable Misstatement"] == 500.0
    assert params["Expected Misstatement"] == 100.0
    assert params["Assurance Factor"] == 4.0
    assert params["Methodology"] == "RSM Random Non-Statistical"

    workbook.close()
