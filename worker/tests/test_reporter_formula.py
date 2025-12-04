"""Ensure High Value Threshold formula is present when override not set."""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from worker.src.models import DataQualityReport, SampleStatistics, SamplingParameters
from worker.src.reporter import generate_reports


def test_high_value_formula_present(tmp_path: Path) -> None:
    sample = []  # Population intentionally empty for formula test

    quality = DataQualityReport(
        total_rows_raw=1,
        total_rows_cleaned=1,
        missing_transaction_id=0,
        missing_amount=0,
        missing_effective_date=0,
        missing_document_type=0,
        missing_description=0,
        invalid_amount_format=0,
        invalid_date_format=0,
        duplicate_transaction_ids=0,
        excluded_due_to_amount=0,
        excluded_due_to_balance=0,
        excluded_zero_amounts=0,
        notes="",
    )
    stats = SampleStatistics(
        population_size=10,
        population_balance_abs=1000.0,
        sampling_interval=100.0,
        high_value_count=0,
        random_sample_count=5,
        coverage_abs=500.0,
        coverage_percent=50.0,
        excluded_zero_amounts=0,
        excluded_due_to_balance=0,
    )
    params = SamplingParameters(
        tolerable_misstatement=1000.0,
        expected_misstatement=100.0,
        assurance_factor=3.0,
        random_seed=1,
    )
    output_path = generate_reports(
        tmp_path,
        sample,
        quality,
        stats,
        params,
        datetime.now(timezone.utc),
        "run-formula",
    )
    wb = load_workbook(output_path, data_only=False)
    sheet = wb["Population Summary"]
    # Find High Value Threshold cell
    for r in range(2, sheet.max_row + 1):
        if sheet.cell(row=r, column=1).value == "High Value Threshold":
            formula_cell = sheet.cell(row=r, column=2)
            assert formula_cell.data_type == "f"
            assert "Parameters Used" in formula_cell.value
            break
    else:
        assert False, "High Value Threshold row not found"
    wb.close()
